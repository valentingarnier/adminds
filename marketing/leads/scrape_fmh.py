"""
Scrape all psychiatrists in Geneva from doctorfmh.ch and export to Excel.

The detail API (phone/email) is rate-limited to ~24 requests per session.
This script rotates sessions via Bright Data Scraping Browser to get fresh
IPs, or falls back to local browser with pauses between batches.

Usage:
    # With Bright Data (recommended — rotates IPs):
    export BRIGHT_DATA_SBR="wss://brd-customer-...:...@brd.superproxy.io:9222"
    python3 leads/scrape_fmh.py

    # Without proxy (will be slow, ~24 profiles per 3-min cycle):
    python3 leads/scrape_fmh.py
"""
import asyncio
import json
import os
import random
import time
from playwright.async_api import async_playwright
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


BASE_URL = "https://doctorfmh.ch"
API_LIST = "/docapi/v1/docs/fr?certs1=39&canton=GE&practice=1&startrecord={offset}"
API_DETAIL = "/docapi/v1/doc/{link_id}/fr"
PAGE_SIZE = 24
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(SCRIPT_DIR, "psychiatrists_geneva.xlsx")
CACHE_FILE = os.path.join(SCRIPT_DIR, "details_cache.json")

# Rate limiting
DETAIL_DELAY = 4.0       # seconds between detail requests
BATCH_SIZE = 20          # profiles per session before rotating
JITTER = 1.0             # random +/- seconds


def load_cache() -> dict:
    if os.path.exists(CACHE_FILE):
        with open(CACHE_FILE, "r") as f:
            cache = json.load(f)
        return cache
    return {}


def save_cache(details: dict):
    with open(CACHE_FILE, "w") as f:
        json.dump(details, f, ensure_ascii=False)


def existing_row_count() -> int:
    if not os.path.exists(OUTPUT_FILE):
        return 0
    try:
        wb = load_workbook(OUTPUT_FILE, read_only=True)
        count = wb.active.max_row - 1
        wb.close()
        return max(count, 0)
    except Exception:
        return 0


async def api_fetch(page, url: str, token: str) -> dict | None:
    try:
        resp_text = await page.evaluate("""async ([url, token]) => {
            const resp = await fetch(url, {
                headers: { 'docfmh': token, 'Accept': 'application/json' }
            });
            return JSON.stringify({status: resp.status, body: await resp.text()});
        }""", [url, token])
        parsed = json.loads(resp_text)
        if parsed["status"] not in (200,):
            return None
        return json.loads(parsed["body"])
    except Exception:
        return None


async def get_session(p, sbr_ws: str | None):
    """Open a browser session and return (browser, page, jwt_token).
    With Bright Data SBR, each call gets a fresh IP."""
    jwt_token = None

    if sbr_ws:
        browser = await p.chromium.connect_over_cdp(sbr_ws)
        context = browser.contexts[0]
        page = await context.new_page()
    else:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
            viewport={"width": 1280, "height": 900},
        )
        page = await context.new_page()

    # Navigate and establish Cloudflare session
    await page.goto(f"{BASE_URL}/fr/", wait_until="domcontentloaded", timeout=120000)
    await page.wait_for_timeout(5000)

    # Accept cookies
    try:
        await page.click("text=Accepter les cookies", timeout=3000)
        await page.wait_for_timeout(500)
    except Exception:
        pass

    # Get JWT token by calling the endpoint directly via fetch
    try:
        jwt_token = await page.evaluate("""async () => {
            // Try to get our public IP first (the API wants it as jtwid)
            let ip = '';
            try {
                const ipResp = await fetch('https://api.ipify.org?format=text');
                ip = await ipResp.text();
            } catch(e) {
                ip = '0.0.0.0';
            }
            const resp = await fetch('/docapi/v1/getjtwtoken?jtwid=' + ip);
            const text = await resp.text();
            return text.trim().replace(/^"|"$/g, '');
        }""")
        if jwt_token and len(jwt_token) > 50:
            pass  # valid token
        else:
            jwt_token = None
    except Exception:
        jwt_token = None
    return browser, page, jwt_token


def export_to_excel(doctors: list[dict], details: dict[str, dict]):
    old_count = existing_row_count()
    if old_count > 0 and len(doctors) < old_count:
        print(f"   SAFETY: Won't overwrite {old_count} rows with {len(doctors)}")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Psychiatres Genève"

    headers = [
        "N°", "Titre", "Prénom", "Nom", "Genre",
        "Lieu de travail", "Lieu de travail 2", "Adresse",
        "Code postal", "Ville", "Canton",
        "EAN/GLN", "Téléphone", "Fax", "Email", "Site web",
        "Langues", "Certifications",
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="6849A0", end_color="6849A0", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for i, doc in enumerate(doctors, 1):
        link_id = doc.get("LINK_ID", "")
        detail = details.get(link_id, {})
        detail_data = detail.get("DATA", {}) if detail else {}
        addresses = detail_data.get("ADDRESSES", []) if detail_data else []

        phone = fax = email = website = languages = certs = ""

        if addresses:
            addr = addresses[0]
            phone = addr.get("PHONE_LINK", "") or ""
            fax = addr.get("FAX_LINK", "") or ""
            email = addr.get("EMAIL_LINK", "") or addr.get("EMAIL", "") or ""
            website = addr.get("WEBSITE_LINK", "") or addr.get("WEBSITE", "") or ""

        langs_list = detail_data.get("LANGUAGES", [])
        if isinstance(langs_list, list):
            languages = ", ".join(l.get("LANGUAGE", "") if isinstance(l, dict) else str(l) for l in langs_list)
        certs_list = detail_data.get("CERTS", [])
        if isinstance(certs_list, list):
            certs = ", ".join(c.get("CERT", "") if isinstance(c, dict) else str(c) for c in certs_list)

        gender = "F" if doc.get("SALUTATION") == "F" else "M" if doc.get("SALUTATION") == "H" else ""

        row_data = [
            i, doc.get("TITLE", ""), doc.get("FIRSTNAME", ""), doc.get("NAME", ""), gender,
            doc.get("AD_NAME", ""), doc.get("AD_NAME_2", ""), doc.get("AD_ADDRESS", ""),
            doc.get("AD_POST_CODE", ""), doc.get("AD_CITY", ""), doc.get("CANTON_CODE", ""),
            doc.get("EAN", ""), phone, fax, email, website, languages, certs,
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=i + 1, column=col, value=value)
            cell.border = thin_border
            if col == 1:
                cell.alignment = Alignment(horizontal="center")

    for col in ws.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

    ws.freeze_panes = "A2"
    wb.save(OUTPUT_FILE)
    print(f"   Excel saved: {OUTPUT_FILE} ({len(doctors)} rows)")


async def main():
    sbr_ws = os.environ.get("BRIGHT_DATA_SBR")
    use_bright = bool(sbr_ws)

    if use_bright:
        print("Mode: Bright Data Scraping Browser (rotating IPs)")
    else:
        print("Mode: Local browser (single IP, slower)")

    async with async_playwright() as p:
        # --- Step 1: Get session and fetch doctor list ---
        print("\n[1/3] Fetching doctor list...")
        browser, page, jwt_token = await get_session(p, sbr_ws)

        if not jwt_token:
            print("   ERROR: No JWT token. Site may be down.")
            await browser.close()
            return

        print(f"   JWT OK — fetching list API...")
        doctors = []
        offset = 0
        total = None

        while True:
            url = BASE_URL + API_LIST.format(offset=offset)
            data = await api_fetch(page, url, jwt_token)
            page_data = (data or {}).get("DATA")

            if not page_data:
                if offset == 0:
                    print("   ERROR: First page returned no data")
                    await browser.close()
                    return
                break

            if total is None:
                total = data.get("RECORDCOUNT", 0)
                print(f"   Total: {total} records")

            doctors.extend(page_data)
            print(f"   Fetched {len(doctors)}/{total}", flush=True)

            if len(doctors) >= total:
                break
            offset += PAGE_SIZE
            await asyncio.sleep(0.3)

        await browser.close()
        print(f"   Done: {len(doctors)} records, {len({d['LINK_ID'] for d in doctors})} unique")

        # --- Step 2: Fetch detail profiles in batches ---
        print(f"\n[2/3] Fetching detail profiles...")
        details = load_cache()
        unique_ids = list({doc["LINK_ID"] for doc in doctors})
        missing = [lid for lid in unique_ids if lid not in details]
        print(f"   {len(unique_ids)} unique, {len(details)} cached, {len(missing)} remaining")

        if missing:
            batch_num = 0
            i = 0
            start_time = time.time()

            while i < len(missing):
                batch_num += 1
                batch = missing[i:i + BATCH_SIZE]
                print(f"\n   --- Batch {batch_num} ({len(batch)} profiles, {i}/{len(missing)} done) ---")

                # Get a fresh session for each batch
                try:
                    browser, page, jwt_token = await get_session(p, sbr_ws)
                except Exception as e:
                    print(f"   Session error: {e}. Retrying in 30s...")
                    await asyncio.sleep(30)
                    continue

                if not jwt_token:
                    print("   No JWT — retrying in 30s...")
                    try:
                        await browser.close()
                    except Exception:
                        pass
                    await asyncio.sleep(30)
                    continue

                successes = 0
                for j, link_id in enumerate(batch):
                    elapsed = time.time() - start_time
                    cached_total = len(details)
                    rate = (cached_total / elapsed * 60) if elapsed > 30 else 0
                    print(
                        f"   [{i+j+1}/{len(missing)}] {link_id[:12]}... "
                        f"({cached_total}/{len(unique_ids)} cached, ~{rate:.0f}/min)",
                        flush=True,
                    )

                    url = BASE_URL + API_DETAIL.format(link_id=link_id)
                    detail = await api_fetch(page, url, jwt_token)

                    if detail and "DATA" in detail:
                        details[link_id] = detail
                        successes += 1
                    else:
                        print(f"   FAIL — stopping batch early")
                        break

                    # Save every few successes
                    if successes % 5 == 0:
                        save_cache(details)

                    await asyncio.sleep(DETAIL_DELAY + random.uniform(-JITTER, JITTER))

                # Save after each batch
                save_cache(details)
                i += successes if successes > 0 else 1  # skip failed ID if 0 successes

                try:
                    await browser.close()
                except Exception:
                    pass

                # Brief pause between batches
                if i < len(missing):
                    pause = 5 if use_bright else 60
                    print(f"   Batch done ({successes}/{len(batch)}). Pausing {pause}s...")
                    await asyncio.sleep(pause)

        # Stats
        with_phone = sum(1 for d in details.values() for a in d.get("DATA", {}).get("ADDRESSES", []) if a.get("PHONE_LINK"))
        with_email = sum(1 for d in details.values() for a in d.get("DATA", {}).get("ADDRESSES", []) if a.get("EMAIL_LINK") or a.get("EMAIL"))
        print(f"\n   Profiles: {len(details)}/{len(unique_ids)}")
        print(f"   With phone: {with_phone}")
        print(f"   With email: {with_email}")

        # --- Step 3: Export ---
        print(f"\n[3/3] Exporting to Excel...")
        export_to_excel(doctors, details)

        print(f"\n{'='*50}")
        print(f"DONE — {len(doctors)} records, {len(details)}/{len(unique_ids)} profiles fetched")
        print(f"Phone: {with_phone} | Email: {with_email}")
        print(f"Cache: {CACHE_FILE}")
        print(f"{'='*50}")


asyncio.run(main())
